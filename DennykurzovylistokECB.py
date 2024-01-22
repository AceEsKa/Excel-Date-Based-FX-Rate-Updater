import pandas as pd
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import datetime

#url
xml = "https://nbs.sk/export/sk/exchange-rate/"

def download(url: str, filename = "exchange-rate.xml"):
    response = requests.get(url)
    
    if response.status_code == 200:
        open(filename,'wb').write(response.content)
    else:
        print(f"Chyba v stahovani dat z {url}")
    

def getEchangeRate(currency = "CZK", filename = "exchange-rate.xml"):
    try:
         with open(filename, 'rb') as file:
            tree = ET.parse(file)
            root = tree.getroot()
    except ET.ParseError as e:
        print(f"Error parsing XML file: {e}")
        exit()
    
    #Define namespaces
    namespaces = {
        'gesmes': 'http://www.gesmes.org/xml/2002-08-01',
        'ecb': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'
    }

    cube_element = root.find(f".//ecb:Cube", namespaces)
    target_cube = cube_element.find(f".//ecb:Cube[@currency='{currency}']", namespaces)
    
    #Extract the CZK rate
    if target_cube is not None:
        exchange_rate = target_cube.attrib.get("rate")
        return exchange_rate
    else:
        return None

def appendToExcel(file_name, column_to, index_since, rates):
    wb = load_workbook(file_name)
    ws = wb['Sheet1']

    colunm_index = (ord(column_to)-64) 
    for index in range(index_since,len(rates)+index_since):
        ws.cell(row = index, column = colunm_index).value = rates[index-index_since]
    
    wb.save(file_name)

def currencyEchangeRatesToExcel():
    file_name = input("Prosim zadajte nazov suboru: ")
    column_from = input("Zadajte stlpec v ktorom sa nachadzaju hodnoty s datumamy: ")
    column_to = input("Zadajte velkym stlpec do ktoreho chcete ulozit kurzy: ")
    index_since = input("Zdajte riadok od ktoreho chcete aby program bral hodnoty: ")
    print("Prosim skontrolujte spravnost zadanych hdonot, ak je vsetko v poriadku a zelate si pokracovat staclte 1 ")
    print("Ak si zelate zadane udaje zmenit prosim stalcte 0")

    while input() != 1:
        file_name = input("Prosim zadajte nazov suboru: ")
        column_from = input("Zadajte stlpec v ktorom sa nachadzaju hodnoty s datumamy: ")
        column_to = input("Zadajte velkym stlpec do ktoreho chcete ulozit kurzy: ")
        index_since = input("Zdajte riadok od ktoreho chcete aby program bral hodnoty: ")
        print("Prosim skontrolujte spravnost zadanych hdonot, ak je vsetko v poriadku a zelate si pokracovat staclte 1 ")
        print("Ak si zelate zadane udaje zmenit prosim stalcte 0")

    print("Program sa po skonceni automaticky vypne")
    
    if ".xlsx" not in file_name:
        file_name += ".xlsx"
    
    data = pd.read_excel(file_name, header=None, usecols=column_from)
    data = data.astype(str)
    
    rates = []
    current = 0 
    previous = 0
    
    for i in range(index_since-1,data.size):
        date = data.iat[i,0]
        
        #we need exchange rate for previous day then the day the transaction happend
        date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        date = datetime.datetime.strptime(date, "%Y-%m-%d")
        date -= datetime.timedelta(days=1)
        date = date.strftime("%Y-%m-%d")
        
        if i!= 0 and date == data.iat[i-1,0]:
            rates.append(previous)
        else:
            download(f"https://nbs.sk/export/sk/exchange-rate/{date}/xml")
            current = getEchangeRate()
            rates.append(current)
            
        previous = current
        
    appendToExcel(file_name, column_to, index_since, rates)

def main():
    print("Prosim uistite sa ze mate vytovrenu kopiu suboru ktory chcete pouizt, na kolko upravy su nezvratne")
    print("Subor ktory si zelate upravit zdajte prosim do priecinka s programom")
    currencyEchangeRatesToExcel()
    
if __name__ == "__main__":
    main()
